import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter

# ── datos ─────────────────────────────────────────────────────────────────────
HEADERS = [
    "N°",
    "CONCEPTO",
    "DEPENDENCIAS",
    "DEPENDIENTES",
    "FÓRMULA ARITMÉTICA ACTUAL",
    "INTERPRETACIÓN CONCEPTUAL",
    "PROPUESTA ARITMÉTICA",
    "PROPUESTA CONCEPTUAL",
]

ROWS = [
    (1,  "Precio Lista Depto",
     "—",
     "Descuento Depto, Precio Lista Total",
     "precioListaDepto (input)",
     "Precio de lista del departamento en UF, proveniente de condiciones comerciales",
     "Sin cambios",
     "Sin cambios"),

    (2,  "Precio Lista Otros (Bienes Conjuntos)",
     "—",
     "Precio Lista Total",
     "SUM(preciosConjuntos[])",
     "Suma de precios lista de estacionamientos y bodegas seleccionadas, en UF. Sin descuento.",
     "Sin cambios",
     "Sin cambios"),

    (3,  "Precio Lista Total",
     "Precio Lista Depto, Precio Lista Otros",
     "Valor de Venta",
     "precioListaDepto + precioListaOtros",
     "Precio de lista bruto del conjunto de unidades cotizadas antes de aplicar descuentos",
     "Sin cambios",
     "Sin cambios"),

    (4,  "Descuento Total (%)",
     "—",
     "Precio Descuento Depto",
     "MIN(descuentoPct + descuentoAdicionalPct, 100%)",
     "Descuento total = condición comercial + adicional negociado, acotado a 100%",
     "Sin cambios",
     "Considerar separar descuento base y adicional como conceptos distintos en el documento para mayor trazabilidad de quién autorizó cada tramo"),

    (5,  "Precio Depto con Descuento",
     "Precio Lista Depto, Descuento Total",
     "Valor de Venta",
     "precioListaDepto × (1 − descTotalPct)",
     "Precio del departamento post-descuento en UF. Descuento aplica SOLO al depto, no a bienes conjuntos.",
     "Sin cambios",
     "Sin cambios"),

    (6,  "Valor de Venta (UF)",
     "Precio Depto con Descuento, Precio Lista Otros",
     "Pie Total, Upfront, Saldo Pie, Cuotón, Pie Construcción, Pie Crédito Directo, Crédito Hip. Base, Plusvalía Año 5",
     "precioDescDepto + precioListaOtros",
     "Precio final de la transacción en UF. Base para casi todos los cálculos posteriores.",
     "Sin cambios",
     "Sin cambios"),

    (7,  "Valor de Venta (CLP)",
     "Valor de Venta (UF), Valor UF",
     "Flujo Mensual, Plusvalía Año 5",
     "valorVentaUF × valorUF",
     "Valor de venta expresado en pesos chilenos usando UF del día",
     "Sin cambios",
     "Sin cambios"),

    (8,  "Pie Total (UF)",
     "Valor de Venta (UF), % Pie",
     "Reserva, Upfront, Saldo Pie, Total Pie Inmobiliaria",
     "valorVentaUF × piePct",
     "Monto total del pie que el comprador debe pagar a la inmobiliaria",
     "Sin cambios",
     "Sin cambios"),

    (9,  "Reserva (UF)",
     "Reserva CLP (input), Valor UF",
     "Saldo Pie",
     "reservaCLP / valorUF",
     "Conversión de la reserva en pesos a UF para descontarla del saldo de pie",
     "Sin cambios",
     "Validar si la reserva puede venir en UF directamente (P4.1 pendiente)"),

    (10, "Upfront Promesa (UF)",
     "Valor de Venta (UF), % Upfront",
     "Saldo Pie",
     "valorVentaUF × upfrontPct",
     "Pago anticipado en la promesa de compraventa, como % del valor de venta",
     "Sin cambios",
     "Sin cambios"),

    (11, "Saldo de Pie (UF)",
     "Pie Total, Reserva, Upfront",
     "Valor Cuota Pie, Total Pie Inmobiliaria",
     "pieTotalUF − reservaUF − upfrontUF",
     "Monto del pie que se paga en cuotas mensuales durante el período de construcción o entrega",
     "Sin cambios",
     "Sin cambios"),

    (12, "Valor Cuota de Pie (UF)",
     "Saldo de Pie, N° Cuotas Pie",
     "—",
     "saldoPieUF / cuotasPieN",
     "Cuota mensual del saldo del pie, distribuida en N cuotas iguales sin interés",
     "Sin cambios",
     "Evaluar si aplica interés en cuotas del pie según contrato (actualmente se asume tasa 0)"),

    (13, "Cuotón (UF)",
     "Valor de Venta (UF), % Cuotón",
     "Total Pie Inmobiliaria",
     "valorVentaUF × cuotonPct",
     "Pago único adicional a la inmobiliaria en promesa o escritura (modalidad especial P3.C2)",
     "Sin cambios",
     "Sin cambios"),

    (14, "Pie Período Construcción (UF)",
     "Valor de Venta (UF), % Pie Construcción",
     "Total Pie Inmobiliaria",
     "valorVentaUF × piePeriodoConstruccionPct",
     "Pie adicional pagado en cuotas decrecientes durante el período de construcción. Se suma al pie estándar.",
     "Sin cambios",
     "Las cuotas decrecientes no están modeladas individualmente; actualmente solo se calcula el monto total. Considerar modelar el calendario de cuotas."),

    (15, "Pie Crédito Directo (UF)",
     "Valor de Venta (UF), % Crédito Directo",
     "— (no afecta CH bancario)",
     "valorVentaUF × pieCreditoDirectoPct",
     "Porcentaje del valor que la inmobiliaria financia directamente al comprador. No reduce el crédito hipotecario bancario.",
     "Sin cambios",
     "Confirmar si el crédito directo genera un plan de pago separado con tasa e interés propio"),

    (16, "Total Pie a Inmobiliaria (UF)",
     "Pie Total, Cuotón, Pie Construcción",
     "Crédito Hipotecario Final",
     "pieTotalUF + cuotonUF + piePeriodoConstruccionUF",
     "Suma total de todos los pagos que el comprador realiza directamente a la inmobiliaria (excluye crédito directo que es financiamiento separado)",
     "Sin cambios",
     "Sin cambios"),

    (17, "Crédito Hip. Base (UF)",
     "Valor de Venta (UF), % Pie",
     "Tasación (cuando sin bono)",
     "valorVentaUF × (1 − piePct)",
     "Monto que el comprador necesita financiar con el banco, antes del aporte inmobiliaria",
     "Sin cambios",
     "Este valor es referencial. El monto efectivo lo determina la tasación."),

    (18, "% Crédito Hip. Puro (D33)",
     "% Pie, % Aporte Inmobiliaria",
     "Tasación",
     "1 − piePct − bonoPiePct",
     "Fracción del precio tasado que queda como crédito hipotecario puro, sin pie ni aporte",
     "Sin cambios",
     "Sin cambios"),

    (19, "Tasación Banco (UF)",
     "Crédito Hip. Base, % CH Puro, % Aporte",
     "Aporte Inmobiliaria (UF), Crédito Hip. Final, Cap Rate",
     "creditoHipBase / (1 − piePct − bonoPiePct) [si bonoPie>0] / valorVentaUF [si bonoPie=0]",
     "Valor que el banco usa como base de la operación. Con aporte inmobiliaria: la tasación 'infla' el valor para que el banco perciba un LTV menor. Sin aporte: tasación = valor de venta.",
     "Sin cambios",
     "Sin cambios"),

    (20, "Aporte Inmobiliaria (UF)",
     "Tasación Banco, % Aporte",
     "—",
     "tasacionUF × bonoPiePct",
     "Monto en UF que la inmobiliaria aporta al banco como parte del precio de compraventa, elevando la tasación y reduciendo el LTV percibido",
     "Sin cambios",
     "Sin cambios"),

    (21, "Crédito Hipotecario Final (UF)",
     "Tasación Banco, Total Pie Inmobiliaria",
     "Cuota Mensual",
     "tasacionUF − totalPieInmobUF",
     "Monto efectivo que el banco presta al comprador. Es la tasación menos todo lo que el comprador ya pagó de pie.",
     "Sin cambios",
     "Verificar si el crédito directo inmobiliaria debe deducirse también del monto bancario en casos específicos"),

    (22, "Plusvalía Acumulada (%)",
     "% Plusvalía Anual",
     "Precio Venta Año 5",
     "(1 + plusvaliaAnual)^5 − 1",
     "Apreciación acumulada del inmueble en 5 años, calculada con interés compuesto",
     "Sin cambios",
     "Sin cambios"),

    (23, "Precio Venta Año 5 (CLP)",
     "Valor de Venta CLP, Plusvalía Acumulada",
     "ROI 5 Años",
     "valorVentaCLP × (1 + plusvaliaAcumulada) × 0.95",
     "Precio estimado de venta al año 5, aplicando plusvalía y un haircut del 5% por costos de venta",
     "Sin cambios",
     "El haircut 0.95 es fijo. Considerar hacerlo parametrizable por inmobiliaria o tipo de proyecto"),

    (24, "Cuota Mensual CH (CLP) [por escenario]",
     "Crédito Hip. Final CLP, CAE, Plazo",
     "Flujo Mensual",
     "PMT(cae/12, plazo×12, creditoHipFinalCLP)",
     "Dividendo mensual del crédito hipotecario calculado con la función PMT de Excel. Una cuota por cada escenario CAE.",
     "Sin cambios",
     "Sin cambios"),

    (25, "Flujo Mensual Neto (CLP) [por escenario]",
     "Arriendo Estimado, Cuota Mensual",
     "Flujo Acumulado, Cap Rate",
     "arriendoCLP − cuotaMensualCLP",
     "Diferencia mensual entre el arriendo que recibiría el inversionista y la cuota hipotecaria que paga",
     "Sin cambios",
     "No incluye gastos comunes, contribuciones ni vacancia real. Considerar agregar parámetros de gastos operacionales para una evaluación más realista"),

    (26, "Flujo Acumulado 5 años (CLP) [por escenario]",
     "Flujo Mensual",
     "ROI 5 Años",
     "flujoMensual × 11 × 5",
     "Flujo neto acumulado en 5 años, asumiendo 11 meses de arriendo por año (1 mes de vacancia)",
     "Sin cambios",
     "El factor 11 meses/año es fijo. Considerar hacerlo parametrizable como \"tasa de vacancia\""),

    (27, "Cap Rate Anual [por escenario]",
     "Arriendo Estimado, Tasación Banco, Valor UF",
     "—",
     "(arriendoCLP × 11 / valorUF) / tasacionUF",
     "Rentabilidad bruta anual del inmueble como activo, expresada como % del valor de tasación. Métrica estándar del mercado inmobiliario.",
     "Sin cambios",
     "No descuenta gastos operacionales. Un Cap Rate neto sería más preciso para comparar proyectos."),

    (28, "ROI 5 Años [por escenario]",
     "Precio Venta Año 5, Valor Venta CLP, Flujo Acumulado, Total Pie Inmobiliaria",
     "ROI Anual",
     "(precioVentaAnio5CLP − valorVentaCLP + flujoAcumuladoCLP) / (totalPieInmobUF × valorUF)",
     "Retorno total sobre la inversión en 5 años, considerando plusvalía y flujo de arriendo neto, como proporción del pie pagado (equity)",
     "Sin cambios",
     "La base del ROI usa solo el pie como equity. Considerar incluir los gastos operacionales y de escrituración para una métrica más conservadora. También evaluar si el crédito directo debe incluirse como parte del equity."),

    (29, "ROI Anual Compuesto [por escenario]",
     "ROI 5 Años",
     "—",
     "(1 + roi5Anios)^(1/5) − 1",
     "Tasa de retorno anual equivalente que produce el ROI a 5 años, calculada con interés compuesto",
     "Sin cambios",
     "Sin cambios"),
]

# ── estilos ───────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F3864")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

EVEN_FILL = PatternFill("solid", fgColor="F2F2F2")
ODD_FILL  = PatternFill("solid", fgColor="FFFFFF")
YLW_FILL  = PatternFill("solid", fgColor="FFFFC0")

DATA_ALIGN_L = Alignment(horizontal="left",   vertical="top", wrap_text=True)
DATA_ALIGN_C = Alignment(horizontal="center", vertical="top", wrap_text=True)

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COL_WIDTHS = [6, 35, 35, 55, 55, 55, 55, 55]

# ── workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Reglas de Cálculo"

# encabezados
for col_idx, hdr in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=hdr)
    cell.fill      = HDR_FILL
    cell.font      = HDR_FONT
    cell.alignment = HDR_ALIGN
    cell.border    = BORDER

ws.row_dimensions[1].height = 30

# datos
for row_data in ROWS:
    r = row_data[0] + 1          # fila en hoja (datos desde fila 2)
    base_fill = EVEN_FILL if row_data[0] % 2 == 0 else ODD_FILL

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=r, column=col_idx, value=value)
        cell.border = BORDER

        # fondo base
        cell.fill = base_fill

        # fondo amarillo en col 7 y 8 si no es "Sin cambios"
        if col_idx == 7 and str(value).strip() != "Sin cambios":
            cell.fill = YLW_FILL
        elif col_idx == 8 and str(value).strip() != "Sin cambios":
            cell.fill = YLW_FILL

        # alineación
        if col_idx == 1:
            cell.alignment = DATA_ALIGN_C
        else:
            cell.alignment = DATA_ALIGN_L

    ws.row_dimensions[r].height = 60

# anchos de columna
for col_idx, width in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# freeze pane
ws.freeze_panes = "A2"

# protección suave (sin contraseña)
ws.protection.sheet = True
ws.protection.password = ""
ws.protection.enable()

# guardar
OUT = r"c:\AI\cotizador-web-mp\RESUMEN_REGLAS_DE_CALCULO.xlsx"
wb.save(OUT)

import os
size = os.path.getsize(OUT)
print(f"Archivo guardado: {OUT}")
print(f"Tamaño: {size:,} bytes ({size/1024:.1f} KB)")
