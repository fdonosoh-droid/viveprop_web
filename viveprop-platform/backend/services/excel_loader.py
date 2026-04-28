"""
Carga los datos desde los archivos .xlsx en data/ y los mantiene en memoria.
"""
from __future__ import annotations
import sys
from pathlib import Path
from typing import Any

import openpyxl

DATA_DIR = Path(__file__).parent.parent.parent / "data"


def _read_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) for h in rows[0]]
    result = []
    for row in rows[1:]:
        if not any(v not in (None, "") for v in row):
            continue
        result.append({headers[i]: row[i] for i in range(len(headers))})
    return result


def _v(val, default=""):
    return default if val is None else val


def _safe_float(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _safe_int(v) -> int:
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


# ─── Loaders ──────────────────────────────────────────────────────────────────

def _load_stock() -> list[dict]:
    path = DATA_DIR / "stock.xlsx"
    if not path.exists():
        print(f"[loader] stock.xlsx no encontrado en {DATA_DIR}", file=sys.stderr)
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = _read_sheet(wb["Stock"])
    wb.close()
    return [{k: ("" if v is None else v) for k, v in r.items()} for r in rows]


def _load_projects() -> list[dict]:
    path = DATA_DIR / "projects.xlsx"
    if not path.exists():
        print(f"[loader] projects.xlsx no encontrado en {DATA_DIR}", file=sys.stderr)
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    proj_rows = _read_sheet(wb["Proyectos"])
    unit_rows = _read_sheet(wb["Unidades"])
    wb.close()

    units_by_proj: dict[str, list] = {}
    for u in unit_rows:
        pid = str(_v(u.get("proyecto_id"), ""))
        units_by_proj.setdefault(pid, []).append(u)

    projects = []
    for p in proj_rows:
        pid = str(_v(p.get("id"), ""))
        amenidades_raw = str(_v(p.get("amenidades"), ""))
        amenidades = [a.strip() for a in amenidades_raw.split(",") if a.strip()]

        units = []
        for u in units_by_proj.get(pid, []):
            units.append({
                "dp":          str(_v(u.get("dp"), "")),
                "tipologia":   str(_v(u.get("tipologia"), "")),
                "piso":        _safe_int(u.get("piso")),
                "m2_interior": _safe_float(u.get("m2_interior")),
                "m2_terraza":  _safe_float(u.get("m2_terraza")),
                "m2_total":    _safe_float(u.get("m2_total")),
                "orientacion": str(_v(u.get("orientacion"), "")),
                "dormitorios": str(_v(u.get("dormitorios"), "")),
                "banos":       str(_v(u.get("banos"), "")),
                "precio_uf":   _safe_float(u.get("precio_uf")),
                "disponible":  bool(_v(u.get("disponible"), True)),
                "estado":      str(_v(u.get("estado"), "")),
            })

        projects.append({
            "id":           pid,
            "nombre":       str(_v(p.get("nombre"), "")),
            "inmobiliaria": str(_v(p.get("inmobiliaria"), "")),
            "comuna":       str(_v(p.get("comuna"), "")),
            "direccion":    str(_v(p.get("direccion"), "")),
            "entrega":      str(_v(p.get("entrega"), "")),
            "descripcion":  str(_v(p.get("descripcion"), "")),
            "foto_portada": str(_v(p.get("foto_portada"), "")),
            "amenidades":   amenidades,
            "unidades":     units,
        })
    return projects


def _load_cc_data() -> dict[str, Any]:
    path = DATA_DIR / "cc_data.xlsx"
    if not path.exists():
        print(f"[loader] cc_data.xlsx no encontrado en {DATA_DIR}", file=sys.stderr)
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = _read_sheet(wb["Condiciones"])
    wb.close()

    meta_cols = {"proyecto_id", "titulo", "nota"}
    campo_labels = [k for k in (rows[0].keys() if rows else []) if k not in meta_cols]

    cc: dict[str, Any] = {}
    for r in rows:
        pid = str(_v(r.get("proyecto_id"), ""))
        if not pid:
            continue
        campos = [
            [label, str(r[label])]
            for label in campo_labels
            if r.get(label) not in (None, "")
        ]
        cc[pid] = {
            "titulo": str(_v(r.get("titulo"), "")),
            "campos": campos,
            "nota":   str(_v(r.get("nota"), "")),
        }
    return cc


def _load_geocodes(filename: str) -> dict[str, list[float]]:
    """Lee geocodes desde el archivo .js legacy (si existe) para no perder el cache."""
    import json, re
    path = DATA_DIR / filename
    if not path.exists():
        return {}
    text = path.read_text(encoding="utf-8")
    lines = [l for l in text.split("\n") if not l.strip().startswith("//")]
    text = "\n".join(lines).strip()
    text = re.sub(r"^\s*const\s+\w+\s*=\s*", "", text)
    if text.rstrip().endswith(";"):
        text = text.rstrip()[:-1]
    try:
        return json.loads(text)
    except Exception:
        return {}


# ─── Singleton ────────────────────────────────────────────────────────────────

class DataStore:
    stock:     list[dict]      = []
    projects:  list[dict]      = []
    cc_data:   dict[str, Any]  = {}
    geocodes:  dict[str, list] = {}
    pri_geocodes: dict[str, list] = {}

    @classmethod
    def load(cls):
        print("[loader] Cargando datos desde Excel...", file=sys.stderr)
        cls.stock     = _load_stock()
        cls.projects  = _load_projects()
        cls.cc_data   = _load_cc_data()
        cls.geocodes      = _load_geocodes("geocodes.js")
        cls.pri_geocodes  = _load_geocodes("pri_geocodes.js")
        print(
            f"[loader] Listo — "
            f"stock:{len(cls.stock)}  "
            f"proyectos:{len(cls.projects)}  "
            f"cc:{len(cls.cc_data)}",
            file=sys.stderr,
        )
