"""
audit_multimedia_pri.py
Genera Excel con estado de imágenes, PDFs y carpetas en disco por proyecto primario.
"""
import sys, json, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT      = Path(__file__).parent.parent
PROJECTS  = ROOT / "frontend" / "public" / "data" / "projects.json"
FOTOS     = ROOT / "frontend" / "public" / "data" / "fotos_pri.json"
PHOTOS    = ROOT / "frontend" / "public" / "photos" / "pri"
OUT       = ROOT / "data" / "audit_multimedia_pri.xlsx"

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
PDF_EXTS   = {".pdf"}

projects = json.loads(PROJECTS.read_text(encoding="utf-8"))
fotos    = json.loads(FOTOS.read_text(encoding="utf-8"))

# Inventario en disco: qué carpetas y portadas existen en photos/pri/
disk_folders = {}   # pid -> Path (o None)
disk_portadas = {}  # pid -> Path (o None)
if PHOTOS.is_dir():
    for item in PHOTOS.iterdir():
        if item.is_dir():
            disk_folders[item.name] = item
        elif item.is_file() and item.suffix.lower() in IMAGE_EXTS:
            # portada: nombre sin extensión = pid
            disk_portadas[item.stem] = item

def disk_folder_info(pid):
    folder = disk_folders.get(pid)
    if folder is None:
        return False, 0, 0, 0   # existe, n_img, n_pdf, n_total
    contents = list(folder.iterdir())
    n_img  = sum(1 for f in contents if f.is_file() and f.suffix.lower() in IMAGE_EXTS)
    n_pdf  = sum(1 for f in contents if f.is_file() and f.suffix.lower() in PDF_EXTS)
    n_total = len(contents)
    return True, n_img, n_pdf, n_total

rows = []
for p in projects:
    pid    = p["id"]
    nombre = p.get("nombre", pid)
    inmob  = p.get("inmobiliaria", "")
    comuna = p.get("comuna", "")
    entrega= p.get("entrega", "")

    unidades      = p.get("unidades", [])
    unidades_disp = sum(1 for u in unidades if u.get("disponible") and u.get("estado","").lower() == "disponible")

    if unidades_disp == 0:
        continue

    # ── Manifest (fotos_pri.json) ─────────────────────────────────────────────
    f = fotos.get(pid, {})
    portada       = f.get("portada", "") or ""
    galeria       = f.get("galeria", []) or []
    pdfs          = f.get("pdfs", []) or []
    tipologias    = f.get("tipologias", []) or []

    tiene_portada_manifest = bool(portada)
    n_galeria_manifest     = len(galeria)
    tiene_galeria_manifest = n_galeria_manifest > 0
    n_pdfs_manifest        = len(pdfs)
    tiene_pdf_manifest     = n_pdfs_manifest > 0

    # ── Disco (photos/pri/) ───────────────────────────────────────────────────
    tiene_portada_disco = pid in disk_portadas
    tiene_carpeta_disco, n_img_disco, n_pdf_disco, n_total_disco = disk_folder_info(pid)

    # ── Estado global ─────────────────────────────────────────────────────────
    faltantes_manifest = []
    if not tiene_portada_manifest:  faltantes_manifest.append("Portada")
    if not tiene_galeria_manifest:  faltantes_manifest.append("Galería")
    if not tiene_pdf_manifest:      faltantes_manifest.append("PDF")
    estado_manifest = "Completo" if not faltantes_manifest else "Falta: " + " + ".join(faltantes_manifest)

    faltantes_disco = []
    if not tiene_portada_disco: faltantes_disco.append("Portada")
    if not tiene_carpeta_disco: faltantes_disco.append("Carpeta")
    elif n_img_disco == 0:      faltantes_disco.append("Fotos en carpeta")
    estado_disco = "OK" if not faltantes_disco else "Falta: " + " + ".join(faltantes_disco)

    rows.append({
        "id":                      pid,
        "nombre":                  nombre,
        "inmobiliaria":            inmob,
        "comuna":                  comuna,
        "entrega":                 entrega,
        "unidades_disp":           unidades_disp,
        # Manifest
        "tiene_portada_manifest":  tiene_portada_manifest,
        "n_galeria":               n_galeria_manifest,
        "tiene_galeria_manifest":  tiene_galeria_manifest,
        "n_pdfs":                  n_pdfs_manifest,
        "tiene_pdf_manifest":      tiene_pdf_manifest,
        "n_tipologias":            len(tipologias),
        "estado_manifest":         estado_manifest,
        # Disco
        "portada_disco":           tiene_portada_disco,
        "carpeta_disco":           tiene_carpeta_disco,
        "n_img_disco":             n_img_disco,
        "n_pdf_disco":             n_pdf_disco,
        "estado_disco":            estado_disco,
    })

# Ordenar: más carencias primero, luego por unidades desc
def sort_key(r):
    m = (0 if r["tiene_portada_manifest"] else 4) + \
        (0 if r["tiene_galeria_manifest"] else 2) + \
        (0 if r["tiene_pdf_manifest"] else 1)
    d = (0 if r["portada_disco"] else 2) + (0 if r["carpeta_disco"] else 1)
    return (-(m + d), -r["unidades_disp"])

rows.sort(key=sort_key)

# ── Estilos ──────────────────────────────────────────────────────────────────
RED    = PatternFill("solid", fgColor="FEE2E2")
AMBER  = PatternFill("solid", fgColor="FEF3C7")
GREEN  = PatternFill("solid", fgColor="DCFCE7")
GRAY   = PatternFill("solid", fgColor="F3F4F6")
HEADER = PatternFill("solid", fgColor="1E3A5F")
HEADER2= PatternFill("solid", fgColor="374151")
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BOLD     = Font(name="Calibri", bold=True, size=10)
NORMAL   = Font(name="Calibri", size=10)
CENTER   = Alignment(horizontal="center", vertical="center")
LEFT     = Alignment(horizontal="left",   vertical="center")
thin     = Side(border_style="thin", color="D1D5DB")
BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)

bool_labels = {True: "Si", False: "No"}

def cell_fill(key, val):
    bool_cols_ok = {"tiene_portada_manifest","tiene_galeria_manifest","tiene_pdf_manifest",
                    "portada_disco","carpeta_disco"}
    if key in bool_cols_ok:
        return GREEN if val else RED
    if key == "estado_manifest":
        if val == "Completo": return GREEN
        if "Portada" in val:  return RED
        return AMBER
    if key == "estado_disco":
        if val == "OK":         return GREEN
        if "Carpeta" in val:    return RED
        return AMBER
    return None

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Auditoria Multimedia"
ws.freeze_panes = "A3"

# Título
ws.merge_cells("A1:S1")
tc = ws["A1"]
tc.value = "AUDITORIA MULTIMEDIA — PROYECTOS MERCADO PRIMARIO"
tc.font  = Font(name="Calibri", bold=True, size=13, color="1E3A5F")
tc.alignment = CENTER
ws.row_dimensions[1].height = 26

# Encabezados — dos grupos: Manifest | Disco
headers = [
    # (label, width, grupo)
    ("ID Proyecto",         16, "base"),
    ("Nombre",              32, "base"),
    ("Inmobiliaria",        18, "base"),
    ("Comuna",              13, "base"),
    ("Entrega",             11, "base"),
    ("Unid. Disp.",         10, "base"),
    # Manifest
    ("Portada (manifest)",  16, "manifest"),
    ("N Fotos Galeria",     14, "manifest"),
    ("Galeria OK",          11, "manifest"),
    ("N PDFs",               8, "manifest"),
    ("PDF OK",               9, "manifest"),
    ("Tipologias",          10, "manifest"),
    ("Estado Manifest",     26, "manifest"),
    # Disco
    ("Portada en disco",    15, "disco"),
    ("Carpeta en disco",    14, "disco"),
    ("Fotos en carpeta",    14, "disco"),
    ("PDFs en carpeta",     13, "disco"),
    ("Estado Disco",        26, "disco"),
]

col_keys = [
    "id","nombre","inmobiliaria","comuna","entrega","unidades_disp",
    "tiene_portada_manifest","n_galeria","tiene_galeria_manifest","n_pdfs","tiene_pdf_manifest","n_tipologias","estado_manifest",
    "portada_disco","carpeta_disco","n_img_disco","n_pdf_disco","estado_disco",
]

for col_idx, (h, w, grupo) in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx, value=h)
    cell.fill      = HEADER if grupo in ("base","manifest") else HEADER2
    cell.font      = HDR_FONT
    cell.alignment = CENTER
    cell.border    = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[2].height = 22

for row_idx, r in enumerate(rows, 3):
    is_gray = row_idx % 2 == 0
    for col_idx, key in enumerate(col_keys, 1):
        val     = r[key]
        display = bool_labels.get(val, val) if isinstance(val, bool) else val
        cell    = ws.cell(row=row_idx, column=col_idx, value=display)
        cell.font   = NORMAL
        cell.border = BORDER
        f = cell_fill(key, r[key])
        if f:           cell.fill = f
        elif is_gray:   cell.fill = GRAY
        cell.alignment = LEFT if key in ("nombre","inmobiliaria","estado_manifest","estado_disco") else CENTER
    ws.row_dimensions[row_idx].height = 15

# ── Hoja resumen ──────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Resumen")
ws2.column_dimensions["A"].width = 46
ws2.column_dimensions["B"].width = 14

ws2.merge_cells("A1:B1")
t2 = ws2["A1"]
t2.value = "RESUMEN EJECUTIVO"
t2.font  = Font(name="Calibri", bold=True, size=12, color="1E3A5F")
t2.alignment = CENTER
ws2.row_dimensions[1].height = 22

totales = [
    ("Total proyectos con stock disponible",              len(rows)),
    ("Proyectos multimedia completos (manifest)",         sum(1 for r in rows if r["estado_manifest"] == "Completo")),
    ("Proyectos con algun faltante (manifest)",           sum(1 for r in rows if r["estado_manifest"] != "Completo")),
    ("--- Detalle manifest ---",                          ""),
    ("  Sin portada",                                     sum(1 for r in rows if not r["tiene_portada_manifest"])),
    ("  Sin galeria de fotos",                            sum(1 for r in rows if not r["tiene_galeria_manifest"])),
    ("  Sin PDF/brochure",                                sum(1 for r in rows if not r["tiene_pdf_manifest"])),
    ("  Sin portada + galeria + PDF (critico)",           sum(1 for r in rows if not r["tiene_portada_manifest"] and not r["tiene_galeria_manifest"] and not r["tiene_pdf_manifest"])),
    ("--- Detalle disco ---",                             ""),
    ("  Sin portada en disco",                            sum(1 for r in rows if not r["portada_disco"])),
    ("  Sin carpeta en disco",                            sum(1 for r in rows if not r["carpeta_disco"])),
    ("  Con carpeta pero sin fotos",                      sum(1 for r in rows if r["carpeta_disco"] and r["n_img_disco"] == 0)),
    ("  Sin carpeta NI portada en disco (sin nada)",      sum(1 for r in rows if not r["carpeta_disco"] and not r["portada_disco"])),
    ("--- Unidades afectadas ---",                        ""),
    ("  Unidades en proyectos sin nada en disco",         sum(r["unidades_disp"] for r in rows if not r["carpeta_disco"] and not r["portada_disco"])),
    ("  Total unidades en stock",                         sum(r["unidades_disp"] for r in rows)),
]

for i, (lbl, val) in enumerate(totales, 2):
    ca = ws2.cell(row=i, column=1, value=lbl)
    cb = ws2.cell(row=i, column=2, value=val)
    is_section = str(lbl).startswith("---")
    ca.font = Font(name="Calibri", bold=is_section, size=10, color="6B7280" if is_section else "000000")
    cb.font = BOLD
    ca.alignment = LEFT; cb.alignment = CENTER
    ca.border = BORDER;  cb.border = BORDER
    if i % 2 == 0 and not is_section:
        ca.fill = GRAY; cb.fill = GRAY

wb.save(OUT)
print(f"Excel guardado: {OUT}")
print(f"Proyectos con stock: {len(rows)}")
print(f"Sin carpeta en disco: {sum(1 for r in rows if not r['carpeta_disco'])}")
print(f"Sin portada en disco: {sum(1 for r in rows if not r['portada_disco'])}")
