"""
fix_euro_portadas.py
Corrige foto_portada en projects.xlsx para los 21 proyectos EURO cuyo img1.png
ya no existe en disco, y limpia las entradas stale del manifest.
Ejecutar UNA SOLA VEZ, luego correr actualizar_primario.bat.
"""
import sys, re, json
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from pathlib import Path
import openpyxl

ROOT     = Path(__file__).parent.parent
PHOTOS   = ROOT / "photos" / "primario" / "EURO(25)"
XLSX     = ROOT / "data" / "projects.xlsx"
MANIFEST = ROOT / "frontend" / "public" / "data" / "fotos_pri.json"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}

RE_TIP = [
    re.compile(r"^\d+D[,\s]",     re.IGNORECASE),
    re.compile(r"ESTUDIO",         re.IGNORECASE),
    re.compile(r"^\d+\s*DORM",    re.IGNORECASE),
    re.compile(r"BANO|BAÑO|BATH", re.IGNORECASE),
]

def is_tipologia(stem):
    return any(r.search(stem) for r in RE_TIP)

def find_portada(folder):
    if not folder.is_dir():
        return None
    imgs = sorted([f for f in folder.iterdir()
                   if f.is_file() and f.suffix.lower() in IMAGE_EXTS])
    # 1) img-000 / img-001
    for f in imgs:
        if re.match(r"^img-0+\d*\.(jpg|jpeg|png|webp)$", f.name, re.IGNORECASE):
            return f.name
    # 2) img1 / img01
    for f in imgs:
        if re.match(r"^img0*1\.(jpg|jpeg|png|webp)$", f.name, re.IGNORECASE):
            return f.name
    # 3) 0.jpg / 0.png
    for f in imgs:
        if re.match(r"^0\.(jpg|jpeg|png|webp)$", f.name, re.IGNORECASE):
            return f.name
    # 4) _01. pattern (ej. guillermo_mann_01.png)
    for f in imgs:
        if re.search(r"_0*1\.(jpg|jpeg|png|webp)$", f.name, re.IGNORECASE) and not is_tipologia(f.stem):
            return f.name
    # 5) primer numerico no-tipologia (1.jpg, 2.jpg, ...)
    num_imgs = sorted(
        [f for f in imgs if re.match(r"^\d+\.(jpg|jpeg|png|webp)$", f.name) and not is_tipologia(f.stem)],
        key=lambda f: int(f.stem)
    )
    if num_imgs:
        return num_imgs[0].name
    # 6) primer no-tipologia
    for f in imgs:
        if not is_tipologia(f.stem):
            return f.name
    return imgs[0].name if imgs else None


def main():
    # 1. Leer xlsx
    wb_ro = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws_ro = wb_ro["Proyectos"]
    rows = list(ws_ro.iter_rows(values_only=True))
    wb_ro.close()
    headers = [str(h).strip() if h else "" for h in rows[0]]
    col_id = headers.index("id")
    col_fp = headers.index("foto_portada")

    to_fix = {}
    for i, row in enumerate(rows[1:], start=2):
        pid = str(row[col_id] or "").strip()
        fp  = str(row[col_fp] or "").strip()
        if "EURO" in fp and ("img1.png" in fp or "img1.jpg" in fp):
            to_fix[pid] = {"row": i, "fp_old": fp}

    print(f"Proyectos a corregir: {len(to_fix)}")
    print()

    # 2. Encontrar portada correcta para cada uno
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Proyectos"]
    changes = []
    for pid, info in to_fix.items():
        fp_old = info["fp_old"]
        folder_path = Path(fp_old).parent
        folder_full = ROOT / folder_path
        portada = find_portada(folder_full)
        if portada is None:
            print(f"  SKIP {pid}: sin imagenes en disco")
            continue
        fp_new = (folder_path / portada).as_posix()
        ws.cell(row=info["row"], column=col_fp + 1).value = fp_new
        changes.append({"pid": pid, "old": fp_old, "new": fp_new})
        print(f"  {pid}")
        print(f"    antes : {fp_old}")
        print(f"    nuevo : {fp_new}")
    wb.save(XLSX)
    print(f"\nXLSX actualizado: {len(changes)} proyectos")

    # 3. Limpiar entradas stale del manifest
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    pids_changed = {c["pid"] for c in changes}
    removed = [pid for pid in pids_changed if pid in manifest]
    for pid in removed:
        del manifest[pid]
    MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Manifest: {len(removed)} entradas stale eliminadas")
    print()
    print("Listo. Ejecuta ahora: actualizar_primario.bat")


if __name__ == "__main__":
    main()
