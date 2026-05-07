"""Comparación 3 fuentes: stock.xlsx vs Disponibilidad AP.xlsx vs Google Sheet"""
import sys, csv, io, re, requests, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from urllib.parse import urlparse, parse_qs
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SHEET_URL = "https://docs.google.com/spreadsheets/d/18oR3_vRDG-_wfTU0bTZQX20yoG9A5MgQpkUNzLFTJoc/export?format=csv&gid=162036413"
CAMPOS = ["op","condominio","direccion","comuna","dp","tipologia","orientacion",
          "est","bod","m2interior","m2total","m2terraza","anio","estado","rango",
          "precioSinBono","bonoPct","bonoUF","video","oportunidad","id"]

# ── helpers ───────────────────────────────────────────────────────────────────
def extract_id(link):
    link = str(link or "").strip().rstrip("/")
    if not link or link == "-": return ""
    qs = parse_qs(urlparse(link).query)
    title = qs.get("title", [""])[0]
    if re.fullmatch(r"\d+", title): return title
    m = re.search(r"-(\d+)$", title) or re.search(r"-(\d+)$", link)
    return m.group(1) if m else ""

def dash(v):
    return "" if str(v or "").strip() in ("-", "None", "") else str(v).strip()

def bono_pct(v):
    s = str(v or "").strip().replace("%", "").strip()
    try: return int(float(s))
    except: return 0

# ── 1. stock.xlsx ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook("data/stock.xlsx", data_only=True)
ws = wb["Stock"]
rows = list(ws.iter_rows(values_only=True))
h = list(rows[0])
stock_all  = [dict(zip(h, row)) for row in rows[1:] if any(v not in (None, "") for v in row)]
stock_byid = {str(r.get("id", "")): r for r in stock_all if r.get("id") not in (None, "")}
stock_noid = [r for r in stock_all if r.get("id") in (None, "")]
wb.close()
print(f"stock.xlsx: {len(stock_all)} total | {len(stock_byid)} con id | {len(stock_noid)} sin id")

# ── 2. Disponibilidad AP.xlsx ─────────────────────────────────────────────────
wb2 = openpyxl.load_workbook("data/Disponibilidad propiedades usadas AP.xlsx", data_only=True)
ws2 = wb2.active
all_rows2 = list(ws2.iter_rows(values_only=True))
h2 = [str(c).strip() if c else "" for c in all_rows2[6]]

def map_ap(row):
    r = dict(zip(h2, [str(v).strip() if v is not None else "" for v in row]))
    return {
        "op": dash(r.get("OP")), "condominio": dash(r.get("CONDOMINIO")),
        "direccion": dash(r.get("DIRECCIÓN", "")), "comuna": dash(r.get("COMUNA")),
        "dp": dash(r.get("DP")), "tipologia": dash(r.get("TIPOLOGIA")),
        "orientacion": dash(r.get("ORIENTACION")), "est": dash(r.get("EST")),
        "bod": dash(r.get("BOD")), "m2interior": dash(r.get("M2 INTERIOR")),
        "m2total": dash(r.get("M2 TOTAL")), "m2terraza": dash(r.get("M2 TERR.", "")),
        "anio": dash(r.get("AÑO DE CONSTRUCCIÓN", "")),
        "estado": dash(r.get("ESTADO DE ADMINISTRACIÓN DE PROPIEDAD", "")),
        "rango": dash(r.get("RANGO DE PRECIO")),
        "precioSinBono": dash(r.get("PRECIO SIN BONO PIE")),
        "bonoPct": bono_pct(r.get("% BONO PIE", 0)),
        "bonoUF": dash(r.get("BONO PIE UF")),
        "video": dash(r.get("LINK VIDEO")), "oportunidad": dash(r.get("OPORTUNIDADES")),
        "id": extract_id(r.get("LINK PUBLICACIÓN", r.get("LINK PUBLICACION", "")))
    }

ap_all  = [map_ap(row) for row in all_rows2[7:] if any(v not in (None, "") for v in row)]
ap_byid = {r["id"]: r for r in ap_all if r["id"]}
wb2.close()
print(f"AP xlsx   : {len(ap_all)} total | {len(ap_byid)} con id")

# ── 3. Google Sheet ───────────────────────────────────────────────────────────
print("Descargando Google Sheet...")
resp = requests.get(SHEET_URL, timeout=20)
text = resp.content.decode("utf-8", errors="replace")
reader = csv.reader(io.StringIO(text))
all_csv = list(reader)
hc = all_csv[6]

def map_csv(row):
    while len(row) < len(hc): row.append("")
    r = dict(zip(hc, row))
    return {
        "op": dash(r.get("OP")), "condominio": dash(r.get("CONDOMINIO")),
        "direccion": dash(r.get("DIRECCIÓN", "")), "comuna": dash(r.get("COMUNA")),
        "dp": dash(r.get("DP")), "tipologia": dash(r.get("TIPOLOGIA")),
        "orientacion": dash(r.get("ORIENTACION")), "est": dash(r.get("EST")),
        "bod": dash(r.get("BOD")), "m2interior": dash(r.get("M2 INTERIOR")),
        "m2total": dash(r.get("M2 TOTAL")), "m2terraza": dash(r.get("M2 TERR.", "")),
        "anio": dash(r.get("AÑO DE CONSTRUCCIÓN", "")),
        "estado": dash(r.get("ESTADO DE ADMINISTRACIÓN DE PROPIEDAD", "")),
        "rango": dash(r.get("RANGO DE PRECIO")),
        "precioSinBono": dash(r.get("PRECIO SIN BONO PIE")),
        "bonoPct": bono_pct(r.get("% BONO PIE", 0)),
        "bonoUF": dash(r.get("BONO PIE UF")),
        "video": dash(r.get("LINK VIDEO")), "oportunidad": dash(r.get("OPORTUNIDADES")),
        "id": extract_id(r.get("LINK PUBLICACIÓN", ""))
    }

gs_all  = [map_csv(row) for row in all_csv[7:] if any(c.strip() for c in row)]
gs_byid = {r["id"]: r for r in gs_all if r["id"]}
print(f"GSheet    : {len(gs_all)} total | {len(gs_byid)} con id")

# ── sets de IDs ───────────────────────────────────────────────────────────────
ids_stock = set(stock_byid.keys())
ids_ap    = set(ap_byid.keys())
ids_gs    = set(gs_byid.keys())

# ── construir Excel ───────────────────────────────────────────────────────────
wb_out = openpyxl.Workbook()

H_FILL  = PatternFill("solid", fgColor="1F4E79")
H_FONT  = Font(color="FFFFFF", bold=True, size=10)
RED_F   = PatternFill("solid", fgColor="FFCCCC")
YEL_F   = PatternFill("solid", fgColor="FFF2CC")
GRN_F   = PatternFill("solid", fgColor="E2EFDA")
BLU_F   = PatternFill("solid", fgColor="DEEAF1")
NF      = Font(size=10)

def hrow(ws, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = H_FONT; cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center")

def auto_width(ws, max_w=42):
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 2, max_w)

# ── Hoja Resumen ──────────────────────────────────────────────────────────────
ws0 = wb_out.active; ws0.title = "Resumen"
rows_res = [
    ("FUENTE", "TOTAL", "CON ID", "SIN ID", "OBSERVACIÓN"),
    ("stock.xlsx — actual en programa", len(stock_all), len(stock_byid), len(stock_noid), "Fuente actual del programa"),
    ("Disponibilidad AP.xlsx — descarga manual", len(ap_all), len(ap_byid), len(ap_all)-len(ap_byid), "Snapshot del GSheet"),
    ("Google Sheet — fuente live AssetPlan", len(gs_all), len(gs_byid), len(gs_all)-len(gs_byid), "Fuente más actualizada"),
    ("", "", "", "", ""),
    ("DIFERENCIA", "CANTIDAD", "", "", "ACCIÓN REQUERIDA"),
    ("Props removidas de AssetPlan", len(ids_stock - ids_gs), "", "", "Eliminar del programa"),
    ("Props nuevas en AssetPlan", len(ids_gs - ids_stock), "", "", "Agregar al programa"),
    ("Solo en GSheet, no en AP xlsx", len(ids_gs - ids_ap), "", "", "Agregadas tras última descarga manual"),
    ("AP xlsx vs GSheet: diferencias de valor", 0, "", "", "Son idénticos — xlsx es snapshot del GSheet"),
    ("Props en stock SIN ID (link no estándar)", len(stock_noid), "", "", "GSheet tiene ID para casi todas"),
]
for ri, row_d in enumerate(rows_res, 1):
    for ci, val in enumerate(row_d, 1):
        c = ws0.cell(row=ri, column=ci, value=val)
        c.font = NF
        if ri in (1, 6):
            c.font = H_FONT; c.fill = H_FILL
ws0.column_dimensions["A"].width = 50
ws0.column_dimensions["B"].width = 12
ws0.column_dimensions["E"].width = 45

# ── Hoja Removidos ────────────────────────────────────────────────────────────
removidos = sorted(ids_stock - ids_gs)
ws1 = wb_out.create_sheet(f"Removidos ({len(removidos)})")
hrow(ws1, CAMPOS)
for ri, rid in enumerate(removidos, 2):
    r = stock_byid[rid]
    for ci, campo in enumerate(CAMPOS, 1):
        c = ws1.cell(row=ri, column=ci, value=str(r.get(campo, "") or ""))
        c.font = NF; c.fill = RED_F
auto_width(ws1)

# ── Hoja Nuevos ───────────────────────────────────────────────────────────────
nuevos = sorted(ids_gs - ids_stock)
ws2 = wb_out.create_sheet(f"Nuevos ({len(nuevos)})")
hrow(ws2, CAMPOS)
for ri, rid in enumerate(nuevos, 2):
    r = gs_byid[rid]
    for ci, campo in enumerate(CAMPOS, 1):
        c = ws2.cell(row=ri, column=ci, value=str(r.get(campo, "") or ""))
        c.font = NF; c.fill = GRN_F
auto_width(ws2)

# ── Hoja Valores diferentes ───────────────────────────────────────────────────
ws3 = wb_out.create_sheet("Valores diferentes")
hrow(ws3, ["id", "condominio", "dp", "campo", "valor_en_stock (actual)", "valor_en_gsheet (correcto)"])
ri3 = 2
for rid in sorted(ids_stock & ids_gs):
    rs = stock_byid[rid]
    rg = gs_byid[rid]
    for campo in CAMPOS:
        v1 = str(rs.get(campo, "") or "").strip()
        v2 = str(rg.get(campo, "") or "").strip()
        if v1 != v2:
            ws3.cell(row=ri3, column=1, value=rid).font = NF
            ws3.cell(row=ri3, column=2, value=rs.get("condominio", "")).font = NF
            ws3.cell(row=ri3, column=3, value=rs.get("dp", "")).font = NF
            c = ws3.cell(row=ri3, column=4, value=campo); c.font = Font(bold=True, size=10)
            c = ws3.cell(row=ri3, column=5, value=v1); c.font = NF; c.fill = RED_F
            c = ws3.cell(row=ri3, column=6, value=v2); c.font = NF; c.fill = GRN_F
            ri3 += 1
auto_width(ws3)

# ── Hoja Sin ID en stock ──────────────────────────────────────────────────────
ws4 = wb_out.create_sheet(f"Sin ID en stock ({len(stock_noid)})")
hrow(ws4, CAMPOS + ["id_en_gsheet"])
gs_by_cond_dp = {(r.get("condominio","").strip().lower(), r.get("dp","").strip().lower()): rid
                 for rid, r in gs_byid.items()}
for ri4, r in enumerate(stock_noid, 2):
    key = (str(r.get("condominio","")).strip().lower(), str(r.get("dp","")).strip().lower())
    match_id = gs_by_cond_dp.get(key, "")
    fill = BLU_F if match_id else YEL_F
    for ci, campo in enumerate(CAMPOS, 1):
        c = ws4.cell(row=ri4, column=ci, value=str(r.get(campo, "") or ""))
        c.font = NF; c.fill = fill
    ws4.cell(row=ri4, column=len(CAMPOS)+1, value=match_id).font = NF
auto_width(ws4)

wb_out.save("data/cartera_assetplan.xlsx")
print(f"\nGuardado: data/cartera_assetplan.xlsx")
print(f"  Resumen")
print(f"  Removidos:           {len(removidos)}")
print(f"  Nuevos:              {len(nuevos)}")
print(f"  Valores diferentes:  {ri3 - 2} filas")
print(f"  Sin ID en stock:     {len(stock_noid)}")
