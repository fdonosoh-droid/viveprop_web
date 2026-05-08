"""
excel_to_js.py — Convierte los .xlsx editados en archivos .js para viveprop-platform.
Correr cada vez que se modifiquen los Excel.

Requiere: pip install openpyxl
Uso:      python tools/excel_to_js.py
"""
import json, sys
from pathlib import Path
from datetime import datetime
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalá con: pip install openpyxl")

DATA_DIR   = Path(__file__).parent.parent / "data"
PUBLIC_DIR = Path(__file__).parent.parent / "frontend" / "public" / "data"
PUBLIC_DIR.mkdir(parents=True, exist_ok=True)


def read_sheet(ws):
    """Devuelve lista de dicts usando la primera fila como encabezados."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) for h in rows[0]]
    result = []
    for row in rows[1:]:
        if not any(v not in (None, "") for v in row):
            continue  # saltar filas vacías
        result.append({headers[i]: row[i] for i in range(len(headers))})
    return result


def val(v, default=""):
    """Devuelve el valor o el default si es None."""
    return default if v is None else v


def to_js(var_name, data, comment=""):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines = [f"// Generado por excel_to_js.py — {ts}"]
    if comment:
        lines.append(f"// {comment}")
    lines.append(f"const {var_name} = {json.dumps(data, ensure_ascii=False, indent=2)};")
    return "\n".join(lines) + "\n"


# ─── STOCK (OBSOLETO — reemplazado por tools/gsheet_to_stock.py) ──────────────

def excel_to_stock():
    # Stock secundario ya no se genera desde Excel.
    # Usar: python tools/gsheet_to_stock.py
    print("  (stock: omitido — usar gsheet_to_stock.py)")

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    rows = read_sheet(wb["Stock"])
    wb.close()

    stock = []
    for r in rows:
        stock.append({k: ("" if v is None else v) for k, v in r.items()})

    out = DATA_DIR / "stock.js"
    out.write_text(to_js("STOCK", stock, f"Propiedades: {len(stock)}"), encoding="utf-8")
    (PUBLIC_DIR / "stock.json").write_text(json.dumps(stock, ensure_ascii=False), encoding="utf-8")
    print(f"  → {out}  ({len(stock)} propiedades)")


# ─── FOTOS MANIFEST ───────────────────────────────────────────────────────────

def merge_fotos_manifest(projects: list) -> None:
    """
    Lee frontend/public/data/fotos_pri.json (si existe) y enriquece cada
    proyecto con foto_portada, fotos, tipologias y pdfs del manifest.
    Se llama justo antes de serializar projects en excel_to_projects().
    """
    manifest_path = PUBLIC_DIR / "fotos_pri.json"
    if not manifest_path.exists():
        return  # manifest no generado aún — no hacer nada
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"  (advertencia: no pude leer fotos_pri.json — {e})")
        return

    merged = 0
    for p in projects:
        entry = manifest.get(p["id"])
        if not entry:
            continue
        p["foto_portada"] = entry.get("portada", p["foto_portada"])
        p["fotos"]        = entry.get("galeria", [])
        p["tipologias"]   = entry.get("tipologias", [])
        p["pdfs"]         = entry.get("pdfs", [])
        merged += 1

    print(f"  → fotos_pri.json: {merged} proyectos enriquecidos con fotos")


# ─── PROJECTS ─────────────────────────────────────────────────────────────────

def _safe_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0

def _safe_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0

def excel_to_projects():
    src = DATA_DIR / "projects.xlsx"
    if not src.exists():
        sys.exit(f"No encontré {src}. Primero corré js_to_excel.py")
    print("Generando projects.js ...")

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    proj_rows = read_sheet(wb["Proyectos"])
    unit_rows = read_sheet(wb["Unidades"])
    wb.close()

    # Agrupar unidades por proyecto
    units_by_proj: dict[str, list] = {}
    for u in unit_rows:
        pid = str(val(u.get("proyecto_id"), ""))
        units_by_proj.setdefault(pid, []).append(u)

    projects = []
    for p in proj_rows:
        pid = str(val(p.get("id"), ""))
        amenidades_raw = str(val(p.get("amenidades"), ""))
        amenidades = [a.strip() for a in amenidades_raw.split(",") if a.strip()]

        units = []
        for u in units_by_proj.get(pid, []):
            units.append({
                "dp":          str(val(u.get("dp"), "")),
                "tipologia":   str(val(u.get("tipologia"), "")),
                "piso":        _safe_int(u.get("piso")),
                "m2_interior": _safe_float(u.get("m2_interior")),
                "m2_terraza":  _safe_float(u.get("m2_terraza")),
                "m2_total":    _safe_float(u.get("m2_total")),
                "orientacion": str(val(u.get("orientacion"), "")),
                "dormitorios": str(val(u.get("dormitorios"), "")),
                "banos":       str(val(u.get("banos"), "")),
                "precio_uf":   _safe_float(u.get("precio_uf")),
                "disponible":  bool(val(u.get("disponible"), True)),
                "estado":      str(val(u.get("estado"), "")),
            })

        projects.append({
            "id":           pid,
            "nombre":       str(val(p.get("nombre"), "")),
            "inmobiliaria": str(val(p.get("inmobiliaria"), "")),
            "comuna":       str(val(p.get("comuna"), "")),
            "direccion":    str(val(p.get("direccion"), "")),
            "entrega":      str(val(p.get("entrega"), "")),
            "descripcion":  str(val(p.get("descripcion"), "")),
            "foto_portada": str(val(p.get("foto_portada"), "")),
            "amenidades":   amenidades,
            "fotos":        [],
            "tipologias":   [],
            "pdfs":         [],
            "unidades":     units,
        })

    merge_fotos_manifest(projects)

    total_u = sum(len(p["unidades"]) for p in projects)
    out = DATA_DIR / "projects.js"
    out.write_text(
        to_js("PROJECTS", projects, f"Proyectos: {len(projects)} | Unidades: {total_u}"),
        encoding="utf-8",
    )
    (PUBLIC_DIR / "projects.json").write_text(json.dumps(projects, ensure_ascii=False), encoding="utf-8")
    print(f"  → {out}  ({len(projects)} proyectos, {total_u} unidades)")


# ─── CC_DATA ──────────────────────────────────────────────────────────────────

def excel_to_cc_data():
    src = DATA_DIR / "cc_data.xlsx"
    if not src.exists():
        sys.exit(f"No encontré {src}. Primero corré js_to_excel.py")
    print("Generando cc_data.js ...")

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    rows = read_sheet(wb["Condiciones"])
    wb.close()

    # Columnas que NO son campos comerciales
    meta_cols = {"proyecto_id", "titulo", "nota"}
    campo_labels = [k for k in (rows[0].keys() if rows else []) if k not in meta_cols]

    cc_data = {}
    for r in rows:
        pid = str(val(r.get("proyecto_id"), ""))
        if not pid:
            continue
        campos = [
            [label, str(r[label])]
            for label in campo_labels
            if r.get(label) not in (None, "")
        ]
        cc_data[pid] = {
            "titulo": str(val(r.get("titulo"), "")),
            "campos": campos,
            "nota":   str(val(r.get("nota"), "")),
        }

    out = DATA_DIR / "cc_data.js"
    out.write_text(
        to_js("CC_DATA", cc_data, f"Proyectos con condiciones: {len(cc_data)}"),
        encoding="utf-8",
    )
    (PUBLIC_DIR / "cc.json").write_text(json.dumps(cc_data, ensure_ascii=False), encoding="utf-8")
    print(f"  → {out}  ({len(cc_data)} proyectos)")


# ─── GEOCODES ─────────────────────────────────────────────────────────────────

def copy_geocodes():
    """Copia geocodes.js y pri_geocodes.js como JSON a public/data/."""
    import re
    for src_name, dst_name in [("geocodes.js", "geocodes.json"), ("pri_geocodes.js", "pri_geocodes.json")]:
        src = DATA_DIR / src_name
        if not src.exists():
            print(f"  (omitido: {src_name} no encontrado)")
            continue
        text = src.read_text(encoding="utf-8")
        text = "\n".join(l for l in text.split("\n") if not l.strip().startswith("//"))
        text = re.sub(r"^\s*const\s+\w+\s*=\s*", "", text.strip())
        if text.rstrip().endswith(";"):
            text = text.rstrip()[:-1]
        data = json.loads(text)
        (PUBLIC_DIR / dst_name).write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        print(f"  → {PUBLIC_DIR / dst_name}  ({len(data)} entradas)")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    excel_to_stock()      # obsoleto — solo imprime aviso
    excel_to_projects()
    excel_to_cc_data()
    copy_geocodes()
    print("\nListo. Archivos JSON en frontend/public/data/ — hacé push para actualizar Vercel.")
