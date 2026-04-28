"""
js_to_excel.py — Bootstrap: convierte los archivos .js en .xlsx editables.
Ejecutar UNA SOLA VEZ para generar los Excel iniciales desde los JS actuales.

Requiere: pip install openpyxl
Uso:      python tools/js_to_excel.py
"""
import json, re, sys
from pathlib import Path
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Falta openpyxl. Instalá con: pip install openpyxl")

DATA_DIR = Path(__file__).parent.parent / "data"

HEADER_FILL  = PatternFill("solid", fgColor="3F4895")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def parse_js(filepath):
    text = Path(filepath).read_text(encoding="utf-8")
    lines = [l for l in text.split("\n") if not l.strip().startswith("//")]
    text = "\n".join(lines).strip()
    text = re.sub(r"^\s*const\s+\w+\s*=\s*", "", text)
    if text.rstrip().endswith(";"):
        text = text.rstrip()[:-1]
    return json.loads(text)


def style_headers(ws):
    for cell in ws[1]:
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def autofit(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


# ─── STOCK ────────────────────────────────────────────────────────────────────

def stock_to_excel():
    print("Leyendo stock.js ...")
    stock = parse_js(DATA_DIR / "stock.js")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"

    cols = ["op", "condominio", "direccion", "comuna", "dp", "tipologia",
            "orientacion", "est", "bod", "m2interior", "m2total", "m2terraza",
            "anio", "estado", "rango", "precioSinBono", "bonoPct", "bonoUF",
            "video", "oportunidad", "id"]

    ws.append(cols)
    for prop in stock:
        ws.append([prop.get(c, "") for c in cols])

    style_headers(ws)
    ws.freeze_panes = "A2"
    autofit(ws)

    out = DATA_DIR / "stock.xlsx"
    wb.save(out)
    print(f"  → {out}  ({len(stock)} propiedades)")


# ─── PROJECTS ─────────────────────────────────────────────────────────────────

def projects_to_excel():
    print("Leyendo projects.js ...")
    projects = parse_js(DATA_DIR / "projects.js")

    wb = openpyxl.Workbook()

    # Hoja 1 — Proyectos
    ws_p = wb.active
    ws_p.title = "Proyectos"
    proj_cols = ["id", "nombre", "inmobiliaria", "comuna", "direccion",
                 "entrega", "descripcion", "foto_portada", "amenidades"]
    ws_p.append(proj_cols)

    # Hoja 2 — Unidades
    ws_u = wb.create_sheet("Unidades")
    unit_cols = ["proyecto_id", "dp", "tipologia", "piso",
                 "m2_interior", "m2_terraza", "m2_total",
                 "orientacion", "dormitorios", "banos",
                 "precio_uf", "disponible", "estado"]
    ws_u.append(unit_cols)

    for proj in projects:
        ws_p.append([
            proj.get("id", ""),
            proj.get("nombre", ""),
            proj.get("inmobiliaria", ""),
            proj.get("comuna", ""),
            proj.get("direccion", ""),
            proj.get("entrega", ""),
            proj.get("descripcion", ""),
            proj.get("foto_portada", ""),
            ", ".join(proj.get("amenidades") or []),
        ])
        for u in proj.get("unidades") or []:
            ws_u.append([
                proj["id"],
                u.get("dp", ""),
                u.get("tipologia", ""),
                u.get("piso", 0),
                u.get("m2_interior", 0),
                u.get("m2_terraza", 0),
                u.get("m2_total", 0),
                u.get("orientacion", ""),
                u.get("dormitorios", ""),
                u.get("banos", ""),
                u.get("precio_uf", 0),
                u.get("disponible", True),
                u.get("estado", ""),
            ])

    for ws in (ws_p, ws_u):
        style_headers(ws)
        ws.freeze_panes = "A2"
        autofit(ws)

    total_u = sum(len(p.get("unidades") or []) for p in projects)
    out = DATA_DIR / "projects.xlsx"
    wb.save(out)
    print(f"  → {out}  ({len(projects)} proyectos, {total_u} unidades)")


# ─── CC_DATA ──────────────────────────────────────────────────────────────────

def cc_data_to_excel():
    print("Leyendo cc_data.js ...")
    cc_data = parse_js(DATA_DIR / "cc_data.js")

    # Recolectar todas las etiquetas de campos (en orden de aparición)
    seen, all_labels = set(), []
    for proj_data in cc_data.values():
        for label, _ in proj_data.get("campos") or []:
            if label not in seen:
                all_labels.append(label)
                seen.add(label)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Condiciones"

    ws.append(["proyecto_id", "titulo"] + all_labels + ["nota"])

    for proj_id, proj_data in cc_data.items():
        campos_dict = dict(proj_data.get("campos") or [])
        row = [proj_id, proj_data.get("titulo", "")]
        row += [campos_dict.get(label, "") for label in all_labels]
        row.append(proj_data.get("nota", ""))
        ws.append(row)

    style_headers(ws)
    ws.freeze_panes = "A2"
    autofit(ws)

    out = DATA_DIR / "cc_data.xlsx"
    wb.save(out)
    print(f"  → {out}  ({len(cc_data)} proyectos con condiciones)")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    stock_to_excel()
    projects_to_excel()
    cc_data_to_excel()
    print("\nListo. Editá los .xlsx y luego corré: python tools/excel_to_js.py")
