"""
Ingevec (ingevec.ecore.cl) - Descarga consolidada de stock
Genera un Excel con todas las unidades (Propiedades Asignadas).

Si da error 401 / redirige a login, obtener nuevas credenciales.
"""

import requests
import re
import sys
import os
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─── CONFIGURACION ────────────────────────────────────────────────────────────
USER       = "agencia@databrokers.cl"
PASSWD     = "Ingevec2025*"
BASE       = "https://ingevec.ecore.cl"
STOCK_URL  = f"{BASE}/com/mantenedor/exe.aspx?idx=5629&com=0&cnt=1&conbar=0&conalt="
OUTPUT_DIR = "C:/AI/stock_inmobiliarias"
# ──────────────────────────────────────────────────────────────────────────────

BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-CL,es;q=0.9",
}


def create_session():
    s = requests.Session()
    s.headers.update(BROWSER_HEADERS)
    return s


def login(s):
    """Authenticate and return session with valid cookies."""
    r = s.get(f"{BASE}/sys/int/log.aspx", timeout=30)
    soup = BeautifulSoup(r.text, "html.parser")

    payload = {
        inp["name"]: inp.get("value", "")
        for inp in soup.find_all("input", {"type": "hidden"})
        if inp.get("name")
    }
    payload["usr"] = USER
    payload["psw"] = PASSWD

    r2 = s.post(
        f"{BASE}/sys/int/log.aspx",
        data=payload,
        allow_redirects=True,
        timeout=30,
        headers={"Referer": f"{BASE}/sys/int/log.aspx"},
    )

    if "log.aspx" in r2.url:
        loc = re.search(
            r'window\.(?:top\.)?location(?:\.href)?\s*=\s*["\']([^"\']+)["\']', r2.text
        )
        if loc:
            redir = loc.group(1)
            if not redir.startswith("http"):
                redir = f"{BASE}/sys/int/{redir}"
            s.get(redir, allow_redirects=True, timeout=30)

    if not s.cookies:
        print("ERROR: Login failed - no session cookie")
        sys.exit(1)

    return s


def get_stock_form(s):
    """Load the stock page and return (html, hidden_fields_dict)."""
    r = s.get(STOCK_URL, allow_redirects=False, timeout=30)

    # Check for redirect to login
    if r.status_code in (301, 302):
        print("ERROR: Redirected to login - session expired")
        sys.exit(1)
    js_redir = re.search(
        r'window\.(?:top\.)?location(?:\.href)?\s*=\s*["\']([^"\']+)["\']', r.text
    )
    if js_redir and "log.aspx" in js_redir.group(1):
        print("ERROR: JS redirect to login - session expired")
        sys.exit(1)

    soup = BeautifulSoup(r.text, "html.parser")
    fields = {
        inp["name"]: inp.get("value", "")
        for inp in soup.find_all("input", {"type": "hidden"})
        if inp.get("name")
    }

    # Get total records from JS init
    m = re.search(r'NotionMantencion_Iniciar\(\s*(\d+)\s*,\s*(\d+)\s*,', r.text)
    if m:
        per_page = int(m.group(1))
        total    = int(m.group(2))
        print(f"  Registros totales: {total}  |  Por pagina: {per_page}")
    else:
        print("  (No se encontro info de paginacion)")

    return r.text, fields


def request_excel_export(s, fields):
    """POST the form with acc=xls and toi=2 (all pages) to get Excel file."""
    payload = dict(fields)
    payload["acc"] = "xls"
    payload["toi"] = "2"   # 2 = all pages

    r = s.post(
        STOCK_URL,
        data=payload,
        timeout=120,
        headers={"Referer": STOCK_URL},
        stream=True,
    )

    content_type = r.headers.get("content-type", "")
    print(f"  Export response: {r.status_code}  Content-Type: {content_type}")
    print(f"  Content-Length: {r.headers.get('content-length','unknown')}")

    return r, content_type


def save_raw_response(content, suffix=""):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{OUTPUT_DIR}/ingevec_raw_{ts}{suffix}"
    with open(fname, "wb") as f:
        f.write(content)
    return fname


def parse_html_table(html):
    """Parse the HTML table of units from exe.aspx response."""
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table", {"id": "tabla"})
    if not table:
        # Try any table with class containing 'tbl' or 'grid'
        table = soup.find("table")

    if not table:
        return [], []

    rows = table.find_all("tr")
    if not rows:
        return [], []

    # First row is header
    headers = [th.get_text(strip=True) for th in rows[0].find_all(["th", "td"])]
    data = []
    for row in rows[1:]:
        cells = [td.get_text(strip=True) for td in row.find_all("td")]
        if cells:
            data.append(cells)
    return headers, data


def build_excel_from_data(headers, rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    fill_h = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font_h = Font(color="FFFFFF", bold=True, size=10)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill_h
        cell.font = font_h
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    for row_num, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=row_num, column=col, value=val)

    wb.save(output_path)


def main():
    print("=" * 60)
    print("  Ingevec - Descarga de Stock (Propiedades Asignadas)")
    print("=" * 60)

    s = create_session()

    print("\n[1/4] Autenticando...")
    login(s)
    print("      OK - sesion iniciada")

    print("\n[2/4] Cargando pagina de stock...")
    html, fields = get_stock_form(s)
    print(f"      Campos del formulario: {list(fields.keys())}")

    print("\n[3/4] Solicitando export Excel (todas las paginas)...")
    r, content_type = request_excel_export(s, fields)

    content = r.content
    print(f"      Bytes recibidos: {len(content)}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"{OUTPUT_DIR}/ingevec_stock_{ts}.xlsx"

    # Determine response type
    is_excel = (
        "excel" in content_type.lower()
        or "spreadsheet" in content_type.lower()
        or "octet" in content_type.lower()
        or content[:4] == b"PK\x03\x04"    # xlsx magic
        or content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # xls magic
    )
    is_html = "html" in content_type.lower() or content[:100].lstrip().startswith(b"<")

    if is_excel:
        print(f"\n[4/4] Guardando Excel directamente...")
        with open(output_path, "wb") as f:
            f.write(content)
        print(f"      Guardado en: {output_path}")

        # Quick check
        try:
            wb = load_workbook(output_path, read_only=True)
            ws = wb.active
            print(f"      Filas: {ws.max_row}  Columnas: {ws.max_column}")
            wb.close()
        except Exception as e:
            print(f"      (No se pudo verificar: {e})")

    elif is_html:
        print(f"\n[4/4] Respuesta HTML - parseando tabla...")
        html_resp = content.decode("utf-8", errors="replace")

        # Save for inspection
        raw = save_raw_response(content, ".html")
        print(f"      HTML guardado en: {raw}")

        headers, rows = parse_html_table(html_resp)
        print(f"      Filas parseadas: {len(rows)}")

        if rows:
            build_excel_from_data(headers, rows, output_path)
            print(f"      Excel guardado en: {output_path}")
        else:
            print("      WARN: No se encontraron datos en la tabla")
            # Show snippet of response
            print(f"      Response snippet: {html_resp[:500]}")
    else:
        raw = save_raw_response(content)
        print(f"\n      Respuesta desconocida ({content_type}), guardada en: {raw}")
        print(f"      Primeros bytes: {content[:20]}")

    print("\n" + "=" * 60)
    print("  Listo!")
    print("=" * 60)


if __name__ == "__main__":
    main()
