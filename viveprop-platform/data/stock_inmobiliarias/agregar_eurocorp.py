"""
Agrega los proyectos de EUROCORP al archivo PROYECTOS.xlsx.
Guarda un backup antes de modificar.
"""

import glob
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROYECTOS_PATH = (
    r"H:\Mi unidad\Viveprop\00 - DDBB_AGENCIA"
    r"\STOCK MERCADO PRIMARIO\Stock\Consolida Stock\PROYECTOS.xlsx"
)

# Usar el Excel de Eurocorp mas reciente
euro_files = sorted(glob.glob(r"C:\AI\stock_inmobiliarias\eurocorp_stock_*.xlsx"), reverse=True)
if not euro_files:
    print("ERROR: No se encontro el archivo eurocorp_stock_*.xlsx")
    raise SystemExit(1)
EUROCORP_PATH = euro_files[0]


def backup(path):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dir_, fname = os.path.split(path)
    name, ext = os.path.splitext(fname)
    backup_path = os.path.join(dir_, f"{name}_backup_{ts}{ext}")
    shutil.copy2(path, backup_path)
    return backup_path


def get_row_style(ws, row_num):
    """Devuelve font, fill, alignment, border de la primera celda de una fila."""
    cell = ws.cell(row=row_num, column=1)
    return cell.font, cell.fill, cell.alignment, cell.border


def copy_style(src_cell, dst_cell):
    if src_cell.font:
        dst_cell.font = src_cell.font.copy()
    if src_cell.fill and src_cell.fill.fill_type != "none":
        dst_cell.fill = src_cell.fill.copy()
    if src_cell.alignment:
        dst_cell.alignment = src_cell.alignment.copy()
    if src_cell.border:
        dst_cell.border = src_cell.border.copy()


def main():
    # --- Obtener proyectos de Eurocorp ---
    print(f"Leyendo proyectos de: {EUROCORP_PATH}")
    wb_euro = load_workbook(EUROCORP_PATH)
    ws_euro = wb_euro["Resumen"]
    euro_projects = []
    for row in range(2, ws_euro.max_row + 1):
        nombre = ws_euro.cell(row=row, column=1).value
        if nombre:
            euro_projects.append(nombre.strip())
    euro_projects.sort()
    print(f"  {len(euro_projects)} proyectos encontrados")

    # --- Backup ---
    bk = backup(PROYECTOS_PATH)
    print(f"\nBackup guardado en:\n  {bk}")

    # --- Abrir PROYECTOS.xlsx ---
    wb = load_workbook(PROYECTOS_PATH)
    ws = wb.active

    # Verificar duplicados
    existing = set()
    for row in range(2, ws.max_row + 1):
        n = ws.cell(row=row, column=2).value
        if n:
            existing.add(n.strip().upper())

    to_add = [p for p in euro_projects if p.upper() not in existing]
    skipped = [p for p in euro_projects if p.upper() in existing]

    if skipped:
        print(f"\nYa existen (se omiten): {skipped}")

    print(f"\nAgregando {len(to_add)} proyectos nuevos...")

    # Tomar estilo de referencia de una fila existente (ultima fila de datos)
    ref_row = ws.max_row

    # Agregar filas nuevas
    start_row = ws.max_row + 1
    for i, nombre in enumerate(to_add):
        r = start_row + i
        ws.cell(row=r, column=1, value="EUROCORP")
        ws.cell(row=r, column=2, value=nombre)
        # Columnas 3-7 quedan en blanco para completar manualmente

        # Copiar estilo de fila de referencia
        for col in range(1, 8):
            src = ws.cell(row=ref_row, column=col)
            dst = ws.cell(row=r, column=col)
            copy_style(src, dst)

        print(f"  + {nombre}")

    # Guardar
    wb.save(PROYECTOS_PATH)
    print(f"\nArchivo guardado: {PROYECTOS_PATH}")
    print(f"Total filas ahora: {ws.max_row - 1} proyectos")
    print(f"\nCampos pendientes de completar manualmente para los {len(to_add)} proyectos nuevos:")
    print("  NEMOTECNICO, COMUNA, DIRECCION, TIPO ENTREGA, PERIODO ENTREGA")


if __name__ == "__main__":
    main()
