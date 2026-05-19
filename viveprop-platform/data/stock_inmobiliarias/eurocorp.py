"""
MobySuite - Descarga consolidada de stock EUROCORP
Genera un Excel con todas las unidades de todos los proyectos.

TOKENS: si el script da error 401, copiar un nuevo cURL desde DevTools
y actualizar BEARER_TOKEN y CAPTCHA_TOKEN abajo.
"""

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import sys
import os

# ─── CONFIGURACION ────────────────────────────────────────────────────────────
BEARER_TOKEN  = "6689707c2313896df9d614792d6dddf407ce4078"
CAPTCHA_TOKEN = ("eyJhbGdvcml0aG0iOiJTSEEtMjU2IiwiY2hhbGxlbmdlIjoiNzAyYzEyNDA4Zjc1NzgwMWJiZGMxYmIxNzU3"
                 "MjljNmI4NTE1MTIzZDAyZDc3OWI5NDliZjRkODZjZDRjYmJiOSIsIm51bWJlciI6ODI1OTgsInNhbHQiOiIwNDQx"
                 "NDEzMjdiZjM2ZWM1ZTYwZDY5NjUiLCJzaWduYXR1cmUiOiIyOWQyMmVhY2MyNzhkY2Y2ODBlZGI5ZWZjY2RjMzY1"
                 "MmVlMTBhNzkyNWMwOTMxOWQyNDQ1ZmE3ZGU0MjU0MTEwIiwidG9vayI6MTQyfQ==")

BASE_URL   = "https://euro-api.mobysuite.com/v1/api"
COMPANY_ID = 1
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
# ──────────────────────────────────────────────────────────────────────────────

HEADERS = {
    "accept": "application/json",
    "authorization": f"Bearer {BEARER_TOKEN}",
    "content-type": "application/json",
    "origin": "https://euro.mobysuite.com",
    "referer": "https://euro.mobysuite.com/",
    "x-mobysuite-captcha": CAPTCHA_TOKEN,
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
}

STATE_COLORS = {
    "DISPONIBLE": "C6EFCE",
    "VENDIDO":    "FFCCCC",
    "RESERVADO":  "FFEB9C",
    "BLOQUEADO":  "D9D9D9",
    "PROMESADO":  "FCE4D6",
}

COLUMNS = [
    ("Proyecto",            "nombreProyecto",        22),
    ("Num. Bien",           "numeroDeBien",           10),
    ("Tipo",                "tipoBien",               14),
    ("Piso",                "piso",                    6),
    ("Orientacion",         "orientacion",            12),
    ("Estado",              "estadoBien",             14),
    ("Tipologia",           "tipologiaDepartamentos", 22),
    ("Dormitorios",         "numeroDeDormitorios",    12),
    ("Banos",               "numeroDeBanos",           8),
    ("Sup. Interior m2",    "superficieInterior",     16),
    ("Sup. Terraza m2",     "superficieTerraza",      15),
    ("Sup. Total m2",       "superficieTotal",        14),
    ("Valor Lista",         "valorLista",             13),
    ("Valor Venta",         "valorVenta",             13),
    ("Precio Inmobiliario", "precioInmobiliario",     20),
    ("Moneda",              "monedaPrincipal",         8),
]


def get_all_projects():
    all_projects = []
    page = 0
    while True:
        resp = requests.post(
            f"{BASE_URL}/projects/company/{COMPANY_ID}?page={page}&size=100",
            headers=HEADERS,
            json={"global": None},
            timeout=30,
        )
        if resp.status_code == 401:
            print("  ERROR 401: Token vencido. Actualiza BEARER_TOKEN y CAPTCHA_TOKEN.")
            sys.exit(1)
        resp.raise_for_status()
        data = resp.json()
        projects = data.get("content", [])
        all_projects.extend(projects)
        total_pages = data.get("totalPages", 1)
        print(f"    Pagina {page + 1}/{total_pages} - {len(projects)} proyectos")
        if page >= total_pages - 1:
            break
        page += 1
    return all_projects


def get_project_assets(project_id):
    resp = requests.get(
        f"{BASE_URL}/assets/project/{project_id}",
        headers=HEADERS,
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def style_header_row(ws, num_cols):
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=10)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18


def write_asset_row(ws, row_num, asset, project):
    estado = asset.get("estadoBien", "")
    fill = PatternFill(
        start_color=STATE_COLORS.get(estado, "FFFFFF"),
        end_color=STATE_COLORS.get(estado, "FFFFFF"),
        fill_type="solid",
    )
    values = {
        "nombreProyecto":         project.get("nombreProyecto"),
        "numeroDeBien":           asset.get("numeroDeBien"),
        "tipoBien":               asset.get("tipoBien"),
        "piso":                   asset.get("piso"),
        "orientacion":            asset.get("orientacion"),
        "estadoBien":             asset.get("estadoBien"),
        "tipologiaDepartamentos": asset.get("tipologiaDepartamentos"),
        "numeroDeDormitorios":    asset.get("numeroDeDormitorios"),
        "numeroDeBanos":          asset.get("numeroDeBanos"),
        "superficieInterior":     asset.get("superficieInterior"),
        "superficieTerraza":      asset.get("superficieTerraza"),
        "superficieTotal":        asset.get("superficieTotal"),
        "valorLista":             asset.get("valorLista"),
        "valorVenta":             asset.get("valorVenta"),
        "precioInmobiliario":     asset.get("precioInmobiliario"),
        "monedaPrincipal":        project.get("monedaPrincipal", "UF"),
    }
    for col, (_, field, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=row_num, column=col, value=values.get(field))
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")


def add_summary_sheet(wb, projects, all_assets_by_project):
    ws = wb.create_sheet(title="Resumen", index=0)
    headers = ["Proyecto", "Total Unidades", "Disponibles", "Vendidas", "Reservadas", "Otros", "Moneda"]
    fill_h = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font_h = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill_h
        cell.font = font_h
        cell.alignment = Alignment(horizontal="center")
    for i, w in enumerate([35, 15, 14, 12, 13, 8, 9], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for row, project in enumerate(projects, 2):
        pid = project["id"]
        assets = all_assets_by_project.get(pid, [])
        estados = [a.get("estadoBien", "") for a in assets]
        ws.cell(row=row, column=1, value=project["nombreProyecto"])
        ws.cell(row=row, column=2, value=len(assets))
        ws.cell(row=row, column=3, value=estados.count("DISPONIBLE"))
        ws.cell(row=row, column=4, value=estados.count("VENDIDO"))
        ws.cell(row=row, column=5, value=estados.count("RESERVADO"))
        ws.cell(row=row, column=6, value=len([e for e in estados if e not in ("DISPONIBLE", "VENDIDO", "RESERVADO")]))
        ws.cell(row=row, column=7, value=project.get("monedaPrincipal", "UF"))


def build_excel(projects, all_assets_by_project, output_path):
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "Consolidado"
    for col, (header, _, width) in enumerate(COLUMNS, 1):
        ws_all.cell(row=1, column=col, value=header)
        ws_all.column_dimensions[ws_all.cell(row=1, column=col).column_letter].width = width
    style_header_row(ws_all, len(COLUMNS))
    ws_all.freeze_panes = "A2"
    row = 2
    for project in projects:
        for asset in all_assets_by_project.get(project["id"], []):
            write_asset_row(ws_all, row, asset, project)
            row += 1
    for project in projects:
        pid = project["id"]
        assets = all_assets_by_project.get(pid, [])
        if not assets:
            continue
        ws_p = wb.create_sheet(title=project["nombreProyecto"][:31])
        for col, (header, _, width) in enumerate(COLUMNS[1:], 1):
            ws_p.cell(row=1, column=col, value=header)
            ws_p.column_dimensions[ws_p.cell(row=1, column=col).column_letter].width = width
        style_header_row(ws_p, len(COLUMNS) - 1)
        ws_p.freeze_panes = "A2"
        for r, asset in enumerate(assets, 2):
            estado = asset.get("estadoBien", "")
            fill = PatternFill(
                start_color=STATE_COLORS.get(estado, "FFFFFF"),
                end_color=STATE_COLORS.get(estado, "FFFFFF"),
                fill_type="solid",
            )
            for col, (_, field, _) in enumerate(COLUMNS[1:], 1):
                value = asset.get(field) if field != "monedaPrincipal" else project.get("monedaPrincipal", "UF")
                cell = ws_p.cell(row=r, column=col, value=value)
                cell.fill = fill
    add_summary_sheet(wb, projects, all_assets_by_project)
    wb.save(output_path)


def main():
    print("  [1/3] Obteniendo proyectos...")
    projects = get_all_projects()
    print(f"        Total: {len(projects)} proyectos")

    print("  [2/3] Descargando unidades...")
    all_assets = {}
    total_units = 0
    for p in projects:
        assets = get_project_assets(p["id"])
        all_assets[p["id"]] = assets
        total_units += len(assets)
        disponibles = sum(1 for a in assets if a.get("estadoBien") == "DISPONIBLE")
        print(f"        {p['nombreProyecto']:<40} {len(assets):>4} unidades  ({disponibles} disp.)")
    print(f"        Total: {total_units} unidades")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"eurocorp_stock_{ts}.xlsx")

    print(f"  [3/3] Generando Excel...")
    build_excel(projects, all_assets, output_path)
    print(f"        Guardado: {output_path}")
    return output_path


if __name__ == "__main__":
    print("=" * 55)
    print("  MobySuite - Stock EUROCORP")
    print("=" * 55)
    main()
    print("=" * 55)
