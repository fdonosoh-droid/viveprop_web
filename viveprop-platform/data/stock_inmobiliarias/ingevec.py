"""
Ingevec (ingevec.ecore.cl) - Descarga consolidada de stock
Genera un Excel con todas las unidades (Propiedades Asignadas).

Si redirige al login, verificar que las credenciales sigan vigentes.
"""

import requests
import re
import sys
import os
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook

# ─── CONFIGURACION ────────────────────────────────────────────────────────────
USER       = "agencia@databrokers.cl"
PASSWD     = "Ingevec2025*"
BASE       = "https://ingevec.ecore.cl"
STOCK_URL  = f"{BASE}/com/mantenedor/exe.aspx?idx=5629&com=0&cnt=1&conbar=0&conalt="
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
# ──────────────────────────────────────────────────────────────────────────────

BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-CL,es;q=0.9",
}


def login(s):
    r = s.get(f"{BASE}/sys/int/log.aspx", timeout=30)
    soup = BeautifulSoup(r.text, "html.parser")
    payload = {
        inp["name"]: inp.get("value", "")
        for inp in soup.find_all("input", {"type": "hidden"})
        if inp.get("name")
    }
    payload["usr"] = USER
    payload["psw"] = PASSWD
    r2 = s.post(
        f"{BASE}/sys/int/log.aspx",
        data=payload,
        allow_redirects=True,
        timeout=30,
        headers={"Referer": f"{BASE}/sys/int/log.aspx"},
    )
    if "log.aspx" in r2.url:
        loc = re.search(
            r'window\.(?:top\.)?location(?:\.href)?\s*=\s*["\']([^"\']+)["\']', r2.text
        )
        if loc:
            redir = loc.group(1)
            if not redir.startswith("http"):
                redir = f"{BASE}/sys/int/{redir}"
            s.get(redir, allow_redirects=True, timeout=30)
    if not s.cookies:
        print("  ERROR: Login fallido - verificar credenciales.")
        sys.exit(1)


def get_form_fields(s):
    r = s.get(STOCK_URL, allow_redirects=False, timeout=30)
    if r.status_code in (301, 302):
        print("  ERROR: Redirigido al login - sesion expirada.")
        sys.exit(1)
    js_redir = re.search(
        r'window\.(?:top\.)?location(?:\.href)?\s*=\s*["\']([^"\']+)["\']', r.text
    )
    if js_redir and "log.aspx" in js_redir.group(1):
        print("  ERROR: Redirigido al login - sesion expirada.")
        sys.exit(1)
    soup = BeautifulSoup(r.text, "html.parser")
    fields = {
        inp["name"]: inp.get("value", "")
        for inp in soup.find_all("input", {"type": "hidden"})
        if inp.get("name")
    }
    m = re.search(r'NotionMantencion_Iniciar\(\s*(\d+)\s*,\s*(\d+)\s*,', r.text)
    if m:
        print(f"        Registros totales: {m.group(2)}  |  Por pagina: {m.group(1)}")
    return fields


def download_excel(s, fields):
    payload = dict(fields)
    payload["acc"] = "xls"
    payload["toi"] = "2"
    r = s.post(
        STOCK_URL,
        data=payload,
        timeout=120,
        headers={"Referer": STOCK_URL},
    )
    content_type = r.headers.get("content-type", "")
    return r.content, content_type


def main():
    s = requests.Session()
    s.headers.update(BROWSER_HEADERS)

    print("  [1/3] Autenticando...")
    login(s)
    print("        OK - sesion iniciada")

    print("  [2/3] Cargando pagina de stock...")
    fields = get_form_fields(s)

    print("  [3/3] Descargando Excel (todas las paginas)...")
    content, content_type = download_excel(s, fields)
    print(f"        Bytes recibidos: {len(content):,}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"ingevec_stock_{ts}.xlsx")

    is_excel = (
        "excel" in content_type.lower()
        or "spreadsheet" in content_type.lower()
        or "octet" in content_type.lower()
        or content[:4] == b"PK\x03\x04"
        or content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    )
    if not is_excel:
        print(f"  ERROR: Respuesta inesperada ({content_type}). Verificar sesion.")
        sys.exit(1)

    with open(output_path, "wb") as f:
        f.write(content)

    try:
        wb = load_workbook(output_path, read_only=True)
        ws = wb.active
        print(f"        Filas: {ws.max_row - 1}  |  Columnas: {ws.max_column}")
        wb.close()
    except Exception:
        pass

    print(f"        Guardado: {output_path}")
    return output_path


if __name__ == "__main__":
    print("=" * 55)
    print("  Ingevec - Stock Propiedades Asignadas")
    print("=" * 55)
    main()
    print("=" * 55)
